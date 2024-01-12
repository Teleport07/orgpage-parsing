import time
from selenium.webdriver.support.ui import WebDriverWait
import openpyxl as openpyxl
import requests as requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
import numpy as np
import ssl

#открываем excel файл, в который будем вносить спаршенные данные
wb = openpyxl.load_workbook('путь к excel файлу')
ws = wb['Лист1']

options = webdriver.ChromeOptions()
options.add_argument(r"--user-data-dir=C:\Users\user_name\AppData\Local\Google\Chrome\User Data")
options.add_experimental_option("detach", True)
driver = webdriver.Chrome(executable_path=r'путь к chromedriver.exe', chrome_options=options)

#открываем рубрику, объекты которой нужно спарсить
driver.get('https://www.orgpage.ru/rossiya/наименование рубрики/')
count = 2
count_pages = 1
wait = WebDriverWait(driver, 500)

#ищем и нажимаем кнопку "Показать все подрубрики"
button = driver.find_element(By.XPATH, '//*[@id="main-wrap"]/div/div[1]/div[4]/div/div[1]/div[2]/a')
button.click()

#создаем список ссылок всех подрубрик в открывшемся попапе
element = driver.find_element(By.XPATH, '//*[@id="rubrick-popup"]/div/div/div[2]/ul')
schools = element.find_elements(By.TAG_NAME, 'li')
list_urls = [i.find_element(By.TAG_NAME, 'a').get_attribute('data-link') for i in schools]

time.sleep(2)

#начинаем пробег по подрубрикам
for education_url2 in list_urls:
    flag = False
    driver.get('https://www.orgpage.ru'+education_url2)
    print('рубрика - '+education_url2)
    time.sleep(2)
    check_strong = driver.find_elements(By.TAG_NAME, 'h2')

    #подрубрики могут подразделяться на дополнительные подрубрики
    #в таком случае нужно снова получить ссылки новых подрубрик и пробежаться по ним в каждой подрубрике уровня повыше
    for i in check_strong:
        strong = i.find_element(By.TAG_NAME, 'strong')
        if strong.text == 'Уточните рубрику':
            flag = True
            break

    #случай, когда есть подрубрики уровня ниже
    if flag:
        button_podrubrica = driver.find_element(By.XPATH, '//*[@id="main-wrap"]/div/div[1]/div[4]/div/div[1]/div[2]/a')
        button_podrubrica.click()
        element_podrubrica = driver.find_element(By.XPATH, '//*[@id="rubrick-popup"]/div/div/div[2]/ul')

        #формируем список ссылок подрубрик уровня ниже
        podrubrica = element_podrubrica.find_elements(By.TAG_NAME, 'li')
        list_podrubrica = [i.find_element(By.TAG_NAME, 'a').get_attribute('data-link') for i in podrubrica]

        #итерация по подрубрикам уровня ниже
        for education_url in list_podrubrica:
            driver.get('https://www.orgpage.ru' + education_url)
            time.sleep(2)

            #в каждой подрубрике есть разбиение на регионы
            try:
                all_regions = driver.find_element(By.XPATH, '//*[@id="main-wrap"]/div/div[1]/div[3]/div/div[1]/div[2]/a')
                all_regions.click()
            except:
                all_regions = driver.find_element(By.XPATH,
                                                  '//*[@id="main-wrap"]/div/div[1]/div[4]/div/div[1]/div[2]/a')
                all_regions.click()
            regions = driver.find_element(By.XPATH, '//*[@id="city-select-popup"]/div/div/div[2]/ul')

            #формируем список ссылок регионов
            regions = regions.find_elements(By.TAG_NAME, 'li')
            regions = [i.find_element(By.TAG_NAME, 'a').get_attribute('data-link') for i in regions]

            #итерируем по регионам
            for reg in regions:
                driver.get('https://www.orgpage.ru'+reg)
                while True:
                    # на странице подгружаем дополнительный контент нажатием на кнопку "Показать еще"
                    # до тех пор, пока она появляется
                    try:
                        element1 = WebDriverWait(driver, 10).until(
                            EC.presence_of_element_located(
                                (By.XPATH, '//*[@id="main-wrap"]/div/div[1]/div[2]/div[4]/div/div[1]/button')))
                        element1.click()
                        time.sleep(np.random.randint(1,3))
                    except:
                        print('нет кнопки Показать еще')
                        with open('рубрика регион.txt', 'w',encoding= 'utf-8') as f:
                            f.write(driver.page_source)
                        break
                while True:
                    print(f'Рубрика {reg}')
                    soup = BeautifulSoup(driver.page_source, "html.parser")
                    url = soup.find_all('div', class_='similar-item__title')

                    for url_n in url:
                        ssilka = url_n.find_next('a').get('href')

                        time.sleep(1)
                        r_url = requests.get(ssilka)
                        if r_url is not None:
                            ws.cell(row=count, column=1).value = ssilka
                            print(url_n.find_next('a').get('href'))
                        soup = BeautifulSoup(r_url.text, "html.parser")
                        title = soup.find('div', class_='company-header__title')

                        #парсим название компании
                        if title is not None:
                            try:
                                ws.cell(row=count, column=2).value = title.find_next('h1').text.strip()
                                print(title.find_next('h1').get_text().strip())
                            except:
                                print('Ошибка с заголовком')
                        telephone = soup.find_all('span', class_='company-information__phone')

                        # парсим контактный телефон компании
                        if telephone is not None:
                            phone = [ph.get_text() for ph in telephone]
                            ws.cell(row=count, column=3).value = ', '.join(phone)
                            print(phone)
                        company = soup.find('div', class_='company-about__text')

                        # парсим описание компании
                        if company is not None:
                            about_company = [comp.get_text() for comp in company.find_all('p')]
                            try:
                                ws.cell(row=count, column=4).value = " ".join(about_company)
                                print(about_company)
                            except:
                                print('не читается')
                        adress = soup.find('div', class_='main-address company-information__address-title')

                        # парсим сокращенный адрес компании
                        if adress is not None:
                            ws.cell(row=count, column=5).value = adress.find_next('span').get_text()
                            print(adress)
                        short_company = soup.find('div', class_='company-short-info')

                        # парсим краткое описание компании
                        if short_company is not None:
                            try:
                                ws.cell(row=count, column=6).value = short_company.find('p', class_='about').get_text()
                            except:
                                print('Проблема с кратким описанием комапании')

                        general_adress = soup.find('div', class_='company-information__address-text')

                        # парсим полный адрес компании
                        if general_adress is not None:
                            ws.cell(row=count, column=7).value = general_adress.text
                            print(general_adress.text)

                        count += 1
                    wb.save('путь к excel файлу')
                    time.sleep(2)
                    count_pages += 1
                    break
    #случай, когда нет подрубрик уровня ниже
    else:
        # в каждой подрубрике есть разбиение на регионы
        time.sleep(2)
        try:
            all_regions = driver.find_element(By.XPATH, '//*[@id="main-wrap"]/div/div[1]/div[4]/div/div[1]/div[2]/a')
            all_regions.click()
        except:
            all_regions = driver.find_element(By.XPATH, '//*[@id="main-wrap"]/div/div[1]/div[3]/div/div[1]/div[2]/a')
            all_regions.click()
        regions = driver.find_element(By.XPATH, '//*[@id="city-select-popup"]/div/div/div[2]/ul')

        #формируем список ссылок регионов
        regions = regions.find_elements(By.TAG_NAME, 'li')
        regions = [i.find_element(By.TAG_NAME, 'a').get_attribute('data-link') for i in regions]

        #итерируем по регионам
        for reg in regions:
            driver.get('https://www.orgpage.ru' + reg)
            while True:
                # на странице подгружаем дополнительный контент нажатием на кнопку "Показать еще"
                # до тех пор, пока она появляется
                try:
                    element1 = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located(
                            (By.XPATH, '//*[@id="main-wrap"]/div/div[1]/div[2]/div[4]/div/div[1]/button')))
                    element1.click()
                    time.sleep(np.random.randint(1, 3))
                except:
                    print('нет кнопки Показать еще')
                    with open('рубрика регион.txt', 'w', encoding='utf-8') as f:
                        f.write(driver.page_source)
                    break
            while True:
                print(f'Рубрика {reg}')
                soup = BeautifulSoup(driver.page_source, "html.parser")
                url = soup.find_all('div', class_='similar-item__title')

                for url_n in url:
                    ssilka = url_n.find_next('a').get('href')
                    time.sleep(1)
                    r_url = requests.get(ssilka)
                    if r_url is not None:
                        ws.cell(row=count, column=1).value = ssilka
                        print(url_n.find_next('a').get('href'))
                    soup = BeautifulSoup(r_url.text, "html.parser")
                    title = soup.find('div', class_='company-header__title')

                    #парсим название компании
                    if title is not None:
                        ws.cell(row=count, column=2).value = title.find_next('h1').text.strip()
                        print(title.find_next('h1').get_text().strip())
                    telephone = soup.find_all('span', class_='company-information__phone')

                    # парсим контактный телефон компании
                    if telephone is not None:
                        phone = [ph.get_text() for ph in telephone]
                        ws.cell(row=count, column=3).value = ', '.join(phone)
                        print(phone)
                    company = soup.find('div', class_='company-about__text')

                    #парсим информацию о компании
                    if company is not None:
                        about_company = [comp.get_text() for comp in company.find_all('p')]
                        try:
                            ws.cell(row=count, column=4).value = " ".join(about_company)
                            print(about_company)
                        except:
                            print('не читается')
                    adress = soup.find('div', class_='main-address company-information__address-title')

                    # парсим сокращенный адрес компании
                    if adress is not None:
                        ws.cell(row=count, column=5).value = adress.find_next('span').get_text()
                        print(adress)
                    short_company = soup.find('div', class_='company-short-info')

                    #парсим краткую информацию о компании
                    if short_company is not None:
                        ws.cell(row=count, column=6).value = short_company.find('p', class_='about').get_text()
                    general_adress = soup.find('div', class_='company-information__address-text')

                    #парсим полный адрес компании
                    if general_adress is not None:
                        ws.cell(row=count, column=7).value = general_adress.text
                        print(general_adress.text)
                    count += 1

                wb.save('путь к файлу excel')
                time.sleep(2)
                count_pages += 1
                break

wb.save('путь к файлу excel')
